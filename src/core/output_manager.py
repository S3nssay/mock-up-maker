import os
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Dict, Any, Tuple
import pandas as pd
import requests
import structlog
from PIL import Image
from core.models import (
    GenerationResult,
    BatchProcessingResult,
    ProcessingStatus,
    ProductData
)


logger = structlog.get_logger()


class OutputManager:
    """Manage output files and directory organization"""

    def __init__(self, base_dir: str = "./product_ads"):
        """Initialize output manager with base directory"""
        self.base_dir = Path(base_dir)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Create directory structure
        self._create_directory_structure()

        # Track outputs
        self.downloaded_images: Dict[str, str] = {}
        self.results: List[GenerationResult] = []
        self.log_file = self.base_dir / "logs" / f"processing_{self.timestamp}.log"

    def _create_directory_structure(self) -> None:
        """Create base directory structure"""
        directories = [
            self.base_dir,
            self.base_dir / "brands",
            self.base_dir / "catalog",
            self.base_dir / "logs"
        ]

        for directory in directories:
            directory.mkdir(parents=True, exist_ok=True)

        logger.info(f"Created output structure at: {self.base_dir}")

    def get_brand_directories(self, brand_name: str) -> Tuple[Path, Path]:
        """Get or create brand-specific directories"""
        brand_dir = self.base_dir / "brands" / self._sanitize_filename(brand_name)
        original_dir = brand_dir / "original"
        ads_dir = brand_dir / "ads"

        # Create directories if they don't exist
        original_dir.mkdir(parents=True, exist_ok=True)
        ads_dir.mkdir(parents=True, exist_ok=True)

        return original_dir, ads_dir

    def save_image(
        self,
        result: GenerationResult,
        image_data: Optional[bytes] = None,
        is_ad_version: bool = False
    ) -> Optional[str]:
        """Save image to appropriate directory"""
        try:
            # Get brand directories
            original_dir, ads_dir = self.get_brand_directories(result.brand_name)

            # Determine save directory
            save_dir = ads_dir if is_ad_version else original_dir

            # Generate filename
            filename = self._generate_filename(
                result.product_name,
                result.brand_name,
                is_ad_version
            )

            filepath = save_dir / filename

            # Download image if URL provided and no data
            if not image_data and result.image_url:
                image_data = self._download_image(result.image_url)

            if image_data:
                # Save image
                with open(filepath, 'wb') as f:
                    f.write(image_data)

                # Verify image is valid
                try:
                    img = Image.open(filepath)
                    img.verify()

                    # Get file size
                    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)

                    logger.info(
                        f"Saved {'ad' if is_ad_version else 'original'} image: {filepath} "
                        f"({file_size_mb:.2f} MB)"
                    )

                    # Update result paths
                    if is_ad_version:
                        result.ad_image_path = str(filepath)
                    else:
                        result.local_image_path = str(filepath)

                    return str(filepath)

                except Exception as e:
                    logger.error(f"Invalid image file: {filepath}")
                    os.remove(filepath)
                    return None

            return None

        except Exception as e:
            logger.error(f"Failed to save image: {str(e)}")
            return None

    def _download_image(self, url: str, max_retries: int = 3) -> Optional[bytes]:
        """Download image from URL with retry logic"""
        for attempt in range(max_retries):
            try:
                response = requests.get(url, timeout=30)
                response.raise_for_status()
                return response.content

            except Exception as e:
                logger.warning(f"Download attempt {attempt + 1} failed: {str(e)}")
                if attempt == max_retries - 1:
                    logger.error(f"Failed to download image after {max_retries} attempts")
                    return None

        return None

    def save_results_excel(
        self,
        results: List[GenerationResult],
        products: List[ProductData],
        filename: Optional[str] = None
    ) -> str:
        """Save processing results to Excel file"""
        try:
            # Prepare filename
            if not filename:
                filename = f"results_{self.timestamp}.xlsx"

            filepath = self.base_dir / "catalog" / filename

            # Create DataFrame from results
            data = []
            for i, result in enumerate(results):
                # Find corresponding product
                product = next(
                    (p for p in products if p.product_name == result.product_name),
                    None
                )

                row = {
                    "row_number": product.row_number if product else i + 1,
                    "product_name": result.product_name,
                    "brand_name": result.brand_name,
                    "product_price": product.product_price if product else "",
                    "product_url": product.product_url if product else "",
                    "prompt": product.prompt if product else "",
                    "generated_image": result.local_image_path or "",
                    "ad_image": result.ad_image_path or "",
                    "processing_time": f"{result.processing_time:.2f}s" if result.processing_time else "",
                    "overlay_time": f"{result.overlay_time:.2f}s" if result.overlay_time else "",
                    "api_cost": f"${result.api_cost:.4f}" if result.api_cost else "",
                    "provider_used": result.provider_used.value if result.provider_used else "",
                    "status": result.status.value,
                    "error_message": result.error_message or "",
                    "timestamp": result.timestamp.strftime("%Y-%m-%d %H:%M:%S")
                }

                # Add image URLs used
                if product:
                    row["images_used"] = ", ".join(product.get_reference_images())
                    row["overlay_applied"] = "Yes" if product.add_product_overlay else "No"

                data.append(row)

            # Create DataFrame and save
            df = pd.DataFrame(data)
            df.to_excel(filepath, index=False, engine='openpyxl')

            # Auto-adjust column widths
            self._adjust_excel_column_widths(filepath)

            logger.info(f"Results saved to: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Failed to save results Excel: {str(e)}")
            raise

    def save_brand_summary(
        self,
        results: List[GenerationResult],
        filename: str = "brand_summary.xlsx"
    ) -> str:
        """Save brand-wise summary to Excel"""
        try:
            filepath = self.base_dir / "catalog" / filename

            # Calculate brand statistics
            brand_stats = {}
            for result in results:
                brand = result.brand_name
                if brand not in brand_stats:
                    brand_stats[brand] = {
                        "Brand": brand,
                        "Total Products": 0,
                        "Successful": 0,
                        "Failed": 0,
                        "Total Cost": 0.0,
                        "Avg Processing Time": []
                    }

                brand_stats[brand]["Total Products"] += 1

                if result.status == ProcessingStatus.SUCCESS:
                    brand_stats[brand]["Successful"] += 1
                    brand_stats[brand]["Total Cost"] += result.api_cost or 0
                    if result.processing_time:
                        brand_stats[brand]["Avg Processing Time"].append(result.processing_time)
                elif result.status == ProcessingStatus.FAILED:
                    brand_stats[brand]["Failed"] += 1

            # Calculate averages
            for brand, stats in brand_stats.items():
                avg_times = stats.pop("Avg Processing Time")
                stats["Avg Processing Time"] = (
                    f"{sum(avg_times) / len(avg_times):.2f}s"
                    if avg_times else "N/A"
                )
                stats["Total Cost"] = f"${stats['Total Cost']:.2f}"
                stats["Success Rate"] = f"{(stats['Successful'] / stats['Total Products'] * 100):.1f}%"

            # Create DataFrame
            df = pd.DataFrame(list(brand_stats.values()))
            df.to_excel(filepath, index=False, engine='openpyxl')

            # Auto-adjust column widths
            self._adjust_excel_column_widths(filepath)

            logger.info(f"Brand summary saved to: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Failed to save brand summary: {str(e)}")
            raise

    def create_batch_result(
        self,
        results: List[GenerationResult],
        total_time: float
    ) -> BatchProcessingResult:
        """Create batch processing result"""
        total_rows = len(results)
        successful = sum(1 for r in results if r.status == ProcessingStatus.SUCCESS)
        failed = sum(1 for r in results if r.status == ProcessingStatus.FAILED)
        skipped = sum(1 for r in results if r.status == ProcessingStatus.SKIPPED)
        total_cost = sum(r.api_cost or 0 for r in results)

        return BatchProcessingResult(
            total_rows=total_rows,
            successful=successful,
            failed=failed,
            skipped=skipped,
            total_cost=total_cost,
            total_time=total_time,
            results=results,
            output_directory=str(self.base_dir),
            results_file=str(self.base_dir / "catalog" / f"results_{self.timestamp}.xlsx")
        )

    def write_log(self, message: str, level: str = "INFO") -> None:
        """Write message to log file"""
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_entry = f"[{timestamp}] [{level}] {message}\n"

            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry)

        except Exception as e:
            logger.error(f"Failed to write to log file: {str(e)}")

    @staticmethod
    def _sanitize_filename(name: str) -> str:
        """Sanitize filename for safe file system usage"""
        # Remove/replace invalid characters
        invalid_chars = '<>:"|?*\\/\n\r\t'
        for char in invalid_chars:
            name = name.replace(char, "_")

        # Limit length
        name = name[:100]

        # Remove trailing dots and spaces
        name = name.strip('. ')

        return name or "unnamed"

    def _generate_filename(
        self,
        product_name: str,
        brand_name: str,
        is_ad: bool = False
    ) -> str:
        """Generate unique filename for image"""
        # Sanitize names
        product_safe = self._sanitize_filename(product_name)
        brand_safe = self._sanitize_filename(brand_name)

        # Add timestamp for uniqueness
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Build filename
        suffix = "_ad" if is_ad else ""
        filename = f"{brand_safe}_{product_safe}{suffix}_{timestamp}.png"

        return filename

    @staticmethod
    def _adjust_excel_column_widths(filepath: Path) -> None:
        """Auto-adjust Excel column widths for better readability"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(filepath)
            ws = wb.active

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)

        except Exception as e:
            # Not critical if this fails
            logger.debug(f"Could not adjust Excel column widths: {str(e)}")

    def cleanup_empty_directories(self) -> None:
        """Remove empty brand directories"""
        try:
            brands_dir = self.base_dir / "brands"
            for brand_dir in brands_dir.iterdir():
                if brand_dir.is_dir():
                    # Check if brand directory has any files
                    has_files = False
                    for subdir in brand_dir.iterdir():
                        if subdir.is_dir() and any(subdir.iterdir()):
                            has_files = True
                            break

                    if not has_files:
                        shutil.rmtree(brand_dir)
                        logger.debug(f"Removed empty brand directory: {brand_dir}")

        except Exception as e:
            logger.warning(f"Error during cleanup: {str(e)}")